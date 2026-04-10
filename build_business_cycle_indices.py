# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "pandas",
#   "numpy",
#   "openpyxl",
#   "plotly"
# ]
# ///

from __future__ import annotations

import argparse
from datetime import date, datetime
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from openpyxl import load_workbook
from plotly.subplots import make_subplots


WINDOW = 8
MIN_PERIODS = 4
SMOOTH_WINDOW = 4
EPSILON = 1e-6

SERIES_SPECS = [
    {
        "key": "gdp",
        "name": "Real GDP",
        "workbook": "5206001_Key_Aggregates.xlsx",
        "label": "Gross domestic product: Chain volume measures ;",
        "series_type": "Seasonally Adjusted",
    },
    {
        "key": "gdp_per_capita",
        "name": "Real GDP per capita",
        "workbook": "5206001_Key_Aggregates.xlsx",
        "label": "GDP per capita: Chain volume measures ;",
        "series_type": "Seasonally Adjusted",
    },
    {
        "key": "dwellings_new_used",
        "name": "Dwellings - new and used",
        "workbook": "5206002_Expenditure_Volume_Measures.xlsx",
        "label": "Private ; Gross fixed capital formation - Dwellings - New and Used ;",
        "series_type": "Seasonally Adjusted",
    },
    {
        "key": "dwellings_alterations",
        "name": "Dwellings - alterations and additions",
        "workbook": "5206002_Expenditure_Volume_Measures.xlsx",
        "label": "Private ; Gross fixed capital formation - Dwellings - Alterations and additions ;",
        "series_type": "Seasonally Adjusted",
    },
    {
        "key": "machinery_new",
        "name": "Machinery and equipment - new",
        "workbook": "5206002_Expenditure_Volume_Measures.xlsx",
        "label": "Private ; Gross fixed capital formation - Machinery and equipment - New ;",
        "series_type": "Seasonally Adjusted",
    },
    {
        "key": "machinery_second_hand",
        "name": "Machinery and equipment - second hand",
        "workbook": "5206002_Expenditure_Volume_Measures.xlsx",
        "label": "Private ; Gross fixed capital formation - Machinery and equipment - Net purchase of second hand assets ;",
        "series_type": "Seasonally Adjusted",
    },
    {
        "key": "purchase_vehicles",
        "name": "Purchase of vehicles",
        "workbook": "5206008_Household_Final_Consumption_Expenditure.xlsx",
        "label": "Purchase of vehicles: Chain volume measures ;",
        "series_type": "Seasonally Adjusted",
    },
    {
        "key": "furnishings",
        "name": "Furnishings and household equipment",
        "workbook": "5206008_Household_Final_Consumption_Expenditure.xlsx",
        "label": "Furnishings and household equipment: Chain volume measures ;",
        "series_type": "Seasonally Adjusted",
    },
    {
        "key": "household_consumption",
        "name": "Household consumption",
        "workbook": "5206002_Expenditure_Volume_Measures.xlsx",
        "label": "Households ; Final consumption expenditure ;",
        "series_type": "Seasonally Adjusted",
    },
    {
        "key": "hours_worked",
        "name": "Hours worked",
        "workbook": "5206001_Key_Aggregates.xlsx",
        "label": "Hours worked: Index ;",
        "series_type": "Seasonally Adjusted",
    },
    {
        "key": "non_dwelling_construction",
        "name": "Non-dwelling construction",
        "workbook": "5206002_Expenditure_Volume_Measures.xlsx",
        "label": "Private ; Gross fixed capital formation - Non-dwelling construction - Total ;",
        "series_type": "Seasonally Adjusted",
    },
    {
        "key": "education_training",
        "name": "Education and training",
        "workbook": "5206006_Industry_GVA.xlsx",
        "label": "Education and training (P) ;",
        "series_type": "Seasonally Adjusted",
    },
    {
        "key": "health_social_assistance",
        "name": "Health care and social assistance",
        "workbook": "5206006_Industry_GVA.xlsx",
        "label": "Health care and social assistance (Q) ;",
        "series_type": "Seasonally Adjusted",
    },
    {
        "key": "public_admin_safety",
        "name": "Public administration and safety",
        "workbook": "5206006_Industry_GVA.xlsx",
        "label": "Public administration and safety (O) ;",
        "series_type": "Seasonally Adjusted",
    },
]

GROUPS = [
    (
        "Leading",
        [
            "dwellings_new_used",
            "dwellings_alterations",
            "machinery_new",
            "machinery_second_hand",
            "purchase_vehicles",
            "furnishings",
        ],
        "#1565c0",
    ),
    (
        "Coincident",
        ["household_consumption", "gdp", "hours_worked", "gdp_per_capita"],
        "#2e7d32",
    ),
    (
        "Lagging",
        [
            "non_dwelling_construction",
            "education_training",
            "health_social_assistance",
            "public_admin_safety",
        ],
        "#c62828",
    ),
]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split())


def convert_abs_date(raw_value: object, frequency: str) -> pd.Timestamp:
    if isinstance(raw_value, datetime):
        timestamp = pd.Timestamp(raw_value)
    elif isinstance(raw_value, date):
        timestamp = pd.Timestamp(raw_value)
    elif isinstance(raw_value, (int, float, np.integer, np.floating)):
        timestamp = pd.Timestamp("1899-12-30") + pd.to_timedelta(float(raw_value), unit="D")
    else:
        timestamp = pd.Timestamp(raw_value)

    normalized_frequency = normalize_text(frequency)
    if normalized_frequency == "Quarter":
        return timestamp.to_period("Q-DEC").to_timestamp(how="end").normalize()
    if normalized_frequency == "Month":
        return timestamp.to_period("M").to_timestamp(how="end").normalize()
    raise ValueError(f"Unsupported ABS frequency {frequency!r}.")


def load_series_batch(workbook_path: Path, specs: list[dict[str, str]]) -> dict[str, pd.Series]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Required workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if "Data1" not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Workbook is missing required sheet 'Data1': {workbook_path}")

    worksheet = workbook["Data1"]
    metadata_rows = list(worksheet.iter_rows(min_row=1, max_row=10, values_only=True))
    if len(metadata_rows) != 10:
        workbook.close()
        raise ValueError(f"Workbook metadata rows are incomplete: {workbook_path}")

    label_row = metadata_rows[0]
    series_type_row = metadata_rows[2]
    frequency_row = metadata_rows[4]

    requested = {
        (normalize_text(spec["label"]), normalize_text(spec["series_type"])): spec for spec in specs
    }
    selected_columns: dict[tuple[str, str], dict[str, object]] = {}

    for column_index in range(2, worksheet.max_column + 1):
        label = normalize_text(label_row[column_index - 1])
        series_type = normalize_text(series_type_row[column_index - 1])
        key = (label, series_type)
        if key not in requested:
            continue
        if key in selected_columns:
            workbook.close()
            raise ValueError(
                f"Duplicate series match for {label!r} ({series_type}) in {workbook_path.name}."
            )
        selected_columns[key] = {
            "spec": requested[key],
            "column_index": column_index,
            "frequency": normalize_text(frequency_row[column_index - 1]),
            "points": [],
        }

    missing = [
        spec for request_key, spec in requested.items() if request_key not in selected_columns
    ]
    if missing:
        workbook.close()
        missing_descriptions = ", ".join(
            f"{spec['label']} ({spec['series_type']})" for spec in missing
        )
        raise ValueError(f"Missing required series in {workbook_path.name}: {missing_descriptions}")

    for row in worksheet.iter_rows(min_row=11, values_only=True):
        raw_date = row[0]
        if raw_date is None:
            continue
        for selected in selected_columns.values():
            raw_value = row[selected["column_index"] - 1]
            if raw_value in (None, ""):
                continue
            timestamp = convert_abs_date(raw_date, str(selected["frequency"]))
            selected["points"].append((timestamp, float(raw_value)))

    workbook.close()

    loaded: dict[str, pd.Series] = {}
    for selected in selected_columns.values():
        if not selected["points"]:
            spec = selected["spec"]
            raise ValueError(
                f"Series {spec['label']} ({spec['series_type']}) in {workbook_path.name} contains no data."
            )
        index = pd.DatetimeIndex([point[0] for point in selected["points"]])
        values = [point[1] for point in selected["points"]]
        series = pd.Series(values, index=index, name=selected["spec"]["name"]).sort_index()
        series = series[~series.index.duplicated(keep="first")]
        loaded[selected["spec"]["key"]] = series

    return loaded


def compute_standardized_growth_gap(level_series: pd.Series) -> pd.Series:
    yearly_growth = level_series.pct_change(4)
    rolling_mean = yearly_growth.rolling(window=WINDOW, min_periods=MIN_PERIODS).mean()
    rolling_std = yearly_growth.rolling(window=WINDOW, min_periods=MIN_PERIODS).std()
    rolling_std = rolling_std.clip(lower=EPSILON)
    signal = (yearly_growth - rolling_mean) / rolling_std
    signal.name = level_series.name
    return signal


def build_balanced_panel(series_map: dict[str, pd.Series]) -> pd.DataFrame:
    if not series_map:
        raise ValueError("Cannot build a balanced panel from an empty series map.")

    valid_starts: list[pd.Timestamp] = []
    valid_ends: list[pd.Timestamp] = []
    for series in series_map.values():
        valid = series.dropna()
        if valid.empty:
            raise ValueError(f"Series {series.name!r} has no valid observations after transformation.")
        valid_starts.append(valid.index.min())
        valid_ends.append(valid.index.max())

    common_start = max(valid_starts)
    common_end = min(valid_ends)
    if common_start > common_end:
        raise ValueError("No overlapping balanced sample exists across the requested series.")

    panel = pd.concat(series_map, axis=1).sort_index()
    panel = panel.loc[(panel.index >= common_start) & (panel.index <= common_end)]
    panel = panel.dropna(how="any")
    if panel.empty:
        raise ValueError("Balanced panel is empty after intersecting date ranges and dropping missing values.")
    return panel


def compute_diffusion_depth_index(component_panel: pd.DataFrame) -> pd.Series:
    diffusion = (component_panel < 0.0).sum(axis=1) / component_panel.shape[1]
    negative_only = component_panel.where(component_panel < 0.0)
    depth = negative_only.mean(axis=1).fillna(0.0)
    raw_composite = -(diffusion * depth)
    smoothed_composite = raw_composite.rolling(window=SMOOTH_WINDOW, min_periods=1).mean()
    return smoothed_composite


def compute_recession_flag(gdp_per_capita: pd.Series) -> tuple[pd.Series, pd.Series]:
    qoq_growth = gdp_per_capita.pct_change(1) * 100.0
    recession_flag = ((qoq_growth < 0.0) & (qoq_growth.shift(1) < 0.0)).astype(int)
    recession_flag.name = "recession_flag"
    return qoq_growth, recession_flag


def build_recession_intervals(recession_flag: pd.Series) -> list[tuple[pd.Timestamp, pd.Timestamp]]:
    intervals: list[tuple[pd.Timestamp, pd.Timestamp]] = []
    sorted_flag = recession_flag.fillna(0).astype(int).sort_index()
    in_recession = False
    start: pd.Timestamp | None = None

    for timestamp, value in sorted_flag.items():
        if value == 1 and not in_recession:
            start = timestamp
            in_recession = True
        elif value == 0 and in_recession and start is not None:
            intervals.append((start, timestamp))
            start = None
            in_recession = False

    if in_recession and start is not None:
        intervals.append((start, sorted_flag.index.max()))

    return intervals


def recession_turning_points(recession_flag: pd.Series) -> list[pd.Timestamp]:
    sorted_flag = recession_flag.fillna(0).astype(int).sort_index()
    changed = sorted_flag.ne(sorted_flag.shift(1)) & sorted_flag.shift(1).notna()
    return list(sorted_flag.index[changed])


def quarter_labels(index: pd.Index) -> np.ndarray:
    periods = pd.PeriodIndex(index, freq="Q-DEC")
    return np.array([str(period) for period in periods], dtype=object).reshape(-1, 1)


def create_figure(
    group_indices: dict[str, pd.Series],
    recession_flag: pd.Series,
    cyclical_share: pd.Series,
) -> go.Figure:
    figure = make_subplots(
        rows=3,
        cols=1,
        shared_xaxes=True,
        vertical_spacing=0.04,
        subplot_titles=["Leading index", "Coincident index", "Lagging index"],
    )

    latest_cyclical_share = cyclical_share.dropna()
    subtitle = (
        "Generate-style standardized YoY growth-gap diffusion-depth composites; "
        "higher values indicate more cyclical downside pressure; "
        "recession shading = two consecutive negative GDP per capita quarters"
    )
    if not latest_cyclical_share.empty:
        subtitle = f"{subtitle}; latest cyclical share = {latest_cyclical_share.iloc[-1]:.1f}% of GDP"

    for row_number, (group_name, _, color) in enumerate(GROUPS, start=1):
        series = group_indices[group_name]
        figure.add_trace(
            go.Scatter(
                x=series.index,
                y=series.values,
                mode="lines",
                name=group_name,
                line={"color": color, "width": 3},
                customdata=quarter_labels(series.index),
                hovertemplate="%{customdata[0]}<br>%{y:.2f}<extra>%{fullData.name}</extra>",
            ),
            row=row_number,
            col=1,
        )
        figure.add_hline(
            y=0.0,
            line={"color": "#9e9e9e", "width": 1, "dash": "dot"},
            row=row_number,
            col=1,
        )

    for start, end in build_recession_intervals(recession_flag):
        figure.add_vrect(
            x0=start,
            x1=end,
            fillcolor="rgba(198, 40, 40, 0.10)",
            line_width=0,
            layer="below",
            row="all",
            col=1,
        )

    for turning_point in recession_turning_points(recession_flag):
        figure.add_vline(
            x=turning_point,
            line={"color": "#616161", "width": 1, "dash": "dash"},
            row="all",
            col=1,
        )

    figure.update_layout(
        title={
            "text": f"Australia Business Cycle Dashboard<br><sup>{subtitle}</sup>",
            "x": 0.5,
        },
        template="plotly_white",
        hovermode="x unified",
        height=950,
        legend={"orientation": "h", "yanchor": "bottom", "y": 1.02, "xanchor": "center", "x": 0.5},
        margin={"l": 70, "r": 30, "t": 120, "b": 60},
    )
    figure.update_xaxes(title_text="Quarter", row=3, col=1)
    figure.update_yaxes(title_text="Index", row=1, col=1)
    figure.update_yaxes(title_text="Index", row=2, col=1)
    figure.update_yaxes(title_text="Index", row=3, col=1)
    return figure


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build fixed-basket leading, coincident, and lagging business-cycle indices for Australia."
    )
    parser.add_argument(
        "--national-accounts-dir",
        type=Path,
        default=Path("abs_workbooks/australian_national_accounts"),
        help="Directory containing the ABS national accounts XLSX files.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("business_cycle_indices.html"),
        help="Output HTML path for the Plotly dashboard.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    national_accounts_dir = args.national_accounts_dir.resolve()

    if not national_accounts_dir.exists():
        raise FileNotFoundError(f"National accounts directory not found: {national_accounts_dir}")

    series_by_key: dict[str, pd.Series] = {}
    specs_by_workbook: dict[str, list[dict[str, str]]] = {}
    for spec in SERIES_SPECS:
        specs_by_workbook.setdefault(spec["workbook"], []).append(spec)

    for workbook_name, specs in specs_by_workbook.items():
        workbook_path = national_accounts_dir / workbook_name
        series_by_key.update(load_series_batch(workbook_path, specs))

    transformed = {key: compute_standardized_growth_gap(series) for key, series in series_by_key.items()}
    group_indices: dict[str, pd.Series] = {}

    for group_name, component_keys, _ in GROUPS:
        component_map = {series_by_key[key].name: transformed[key] for key in component_keys}
        component_panel = build_balanced_panel(component_map)
        composite = compute_diffusion_depth_index(component_panel)
        composite.name = group_name
        group_indices[group_name] = composite

    gdp_qoq_growth, recession_flag = compute_recession_flag(series_by_key["gdp_per_capita"])
    if recession_flag.empty:
        raise ValueError("GDP per capita recession flag is empty after calculation.")

    cyclical_components = pd.concat(
        [series_by_key[key].rename(key) for key in GROUPS[0][1]],
        axis=1,
        join="inner",
    ).dropna(how="any")
    if cyclical_components.empty:
        raise ValueError("Cyclical GDP basket has no common observations.")
    cyclical_gdp = cyclical_components.sum(axis=1).rename("cyclical_gdp")

    cyclical_alignment = pd.concat(
        [cyclical_gdp, series_by_key["gdp"].rename("gdp")],
        axis=1,
        join="inner",
    ).dropna(how="any")
    if cyclical_alignment.empty:
        raise ValueError("Cyclical GDP share is empty after aligning cyclical GDP with GDP.")
    cyclical_share = (cyclical_alignment["cyclical_gdp"] / cyclical_alignment["gdp"]) * 100.0

    figure = create_figure(group_indices, recession_flag, cyclical_share)

    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.write_html(output_path, include_plotlyjs=True, full_html=True)

    print(f"Wrote {output_path}")
    print(
        "Latest values: "
        + ", ".join(f"{group}={series.iloc[-1]:.2f}" for group, series in group_indices.items())
    )
    print(f"Latest cyclical share = {cyclical_share.iloc[-1]:.2f}% of GDP")
    print(f"Latest GDP per capita QoQ growth = {gdp_qoq_growth.dropna().iloc[-1]:.2f}%")


if __name__ == "__main__":
    main()
