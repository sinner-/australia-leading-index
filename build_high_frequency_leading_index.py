# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "pandas",
#   "numpy",
#   "openpyxl",
#   "plotly"
# ]
# ///

from __future__ import annotations

import argparse
from datetime import date, datetime
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from openpyxl import load_workbook


WINDOW = 24
MIN_PERIODS = 12
SMOOTH_WINDOW = 4
EPSILON = 1e-6

SERIES_SPECS = [
    {
        "key": "clothing_footwear",
        "name": "Clothing and footwear",
        "label": "Household spending ; Clothing and footwear ; Australia ; Current Price ;",
        "series_type": "Seasonally Adjusted",
    },
    {
        "key": "furnishings",
        "name": "Furnishings and household equipment",
        "label": "Household spending ; Furnishings and household equipment ; Australia ; Current Price ;",
        "series_type": "Seasonally Adjusted",
    },
    {
        "key": "transport",
        "name": "Transport",
        "label": "Household spending ; Transport ; Australia ; Current Price ;",
        "series_type": "Seasonally Adjusted",
    },
]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split())


def convert_abs_date(raw_value: object, frequency: str) -> pd.Timestamp:
    if isinstance(raw_value, datetime):
        timestamp = pd.Timestamp(raw_value)
    elif isinstance(raw_value, date):
        timestamp = pd.Timestamp(raw_value)
    elif isinstance(raw_value, (int, float, np.integer, np.floating)):
        timestamp = pd.Timestamp("1899-12-30") + pd.to_timedelta(float(raw_value), unit="D")
    else:
        timestamp = pd.Timestamp(raw_value)

    normalized_frequency = normalize_text(frequency)
    if normalized_frequency != "Month":
        raise ValueError(f"Expected monthly ABS data, found {frequency!r}.")
    return timestamp.to_period("M").to_timestamp(how="end").normalize()


def load_series(workbook_path: Path) -> dict[str, pd.Series]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Required workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if "Data1" not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Workbook is missing required sheet 'Data1': {workbook_path}")

    worksheet = workbook["Data1"]
    metadata_rows = list(worksheet.iter_rows(min_row=1, max_row=10, values_only=True))
    if len(metadata_rows) != 10:
        workbook.close()
        raise ValueError(f"Workbook metadata rows are incomplete: {workbook_path}")

    label_row = metadata_rows[0]
    series_type_row = metadata_rows[2]
    frequency_row = metadata_rows[4]

    requested = {
        (normalize_text(spec["label"]), normalize_text(spec["series_type"])): spec
        for spec in SERIES_SPECS
    }
    selected_columns: dict[tuple[str, str], dict[str, object]] = {}

    for column_index in range(2, worksheet.max_column + 1):
        label = normalize_text(label_row[column_index - 1])
        series_type = normalize_text(series_type_row[column_index - 1])
        key = (label, series_type)
        if key not in requested:
            continue
        if key in selected_columns:
            workbook.close()
            raise ValueError(
                f"Duplicate series match for {label!r} ({series_type}) in {workbook_path.name}."
            )
        selected_columns[key] = {
            "spec": requested[key],
            "column_index": column_index,
            "frequency": normalize_text(frequency_row[column_index - 1]),
            "points": [],
        }

    missing = [
        spec for request_key, spec in requested.items() if request_key not in selected_columns
    ]
    if missing:
        workbook.close()
        missing_descriptions = ", ".join(
            f"{spec['label']} ({spec['series_type']})" for spec in missing
        )
        raise ValueError(f"Missing required series in {workbook_path.name}: {missing_descriptions}")

    for row in worksheet.iter_rows(min_row=11, values_only=True):
        raw_date = row[0]
        if raw_date is None:
            continue
        for selected in selected_columns.values():
            raw_value = row[selected["column_index"] - 1]
            if raw_value in (None, ""):
                continue
            timestamp = convert_abs_date(raw_date, str(selected["frequency"]))
            selected["points"].append((timestamp, float(raw_value)))

    workbook.close()

    loaded: dict[str, pd.Series] = {}
    for selected in selected_columns.values():
        if not selected["points"]:
            spec = selected["spec"]
            raise ValueError(
                f"Series {spec['label']} ({spec['series_type']}) in {workbook_path.name} contains no data."
            )
        index = pd.DatetimeIndex([point[0] for point in selected["points"]])
        values = [point[1] for point in selected["points"]]
        series = pd.Series(values, index=index, name=selected["spec"]["name"]).sort_index()
        series = series[~series.index.duplicated(keep="first")]
        loaded[selected["spec"]["key"]] = series

    return loaded


def compute_standardized_growth_gap(level_series: pd.Series) -> pd.Series:
    yearly_growth = level_series.pct_change(12)
    rolling_mean = yearly_growth.rolling(window=WINDOW, min_periods=MIN_PERIODS).mean()
    rolling_std = yearly_growth.rolling(window=WINDOW, min_periods=MIN_PERIODS).std()
    rolling_std = rolling_std.clip(lower=EPSILON)
    signal = (yearly_growth - rolling_mean) / rolling_std
    signal.name = level_series.name
    return signal


def build_balanced_panel(series_map: dict[str, pd.Series]) -> pd.DataFrame:
    if not series_map:
        raise ValueError("Cannot build a balanced panel from an empty series map.")

    valid_starts: list[pd.Timestamp] = []
    valid_ends: list[pd.Timestamp] = []
    for series in series_map.values():
        valid = series.dropna()
        if valid.empty:
            raise ValueError(f"Series {series.name!r} has no valid observations after transformation.")
        valid_starts.append(valid.index.min())
        valid_ends.append(valid.index.max())

    common_start = max(valid_starts)
    common_end = min(valid_ends)
    if common_start > common_end:
        raise ValueError("No overlapping balanced sample exists across the requested series.")

    panel = pd.concat(series_map, axis=1).sort_index()
    panel = panel.loc[(panel.index >= common_start) & (panel.index <= common_end)]
    panel = panel.dropna(how="any")
    if panel.empty:
        raise ValueError("Balanced panel is empty after intersecting date ranges and dropping missing values.")
    return panel


def compute_diffusion_depth_index(component_panel: pd.DataFrame) -> pd.Series:
    diffusion = (component_panel < 0.0).sum(axis=1) / component_panel.shape[1]
    negative_only = component_panel.where(component_panel < 0.0)
    depth = negative_only.mean(axis=1).fillna(0.0)
    raw_composite = -(diffusion * depth)
    smoothed_composite = raw_composite.rolling(window=SMOOTH_WINDOW, min_periods=1).mean()
    return smoothed_composite


def month_labels(index: pd.Index) -> np.ndarray:
    return np.array([timestamp.strftime("%Y-%m") for timestamp in index], dtype=object).reshape(-1, 1)


def create_figure(composite: pd.Series) -> go.Figure:
    figure = go.Figure()
    figure.add_trace(
        go.Scatter(
            x=composite.index,
            y=composite.values,
            mode="lines",
            name="Leading index",
            line={"width": 3, "color": "#1b1f23"},
            customdata=month_labels(composite.index),
            hovertemplate="%{customdata[0]}<br>%{y:.2f}<extra>%{fullData.name}</extra>",
        )
    )
    figure.add_hline(y=0.0, line={"color": "#9e9e9e", "width": 1, "dash": "dot"})
    figure.update_layout(
        title={
            "text": (
                "Australia High-Frequency Leading Index"
                "<br><sup>Generate-style standardized YoY growth-gap diffusion-depth composite; "
                "higher values indicate more downside pressure</sup>"
            ),
            "x": 0.5,
        },
        template="plotly_white",
        hovermode="x unified",
        height=560,
        showlegend=False,
        margin={"l": 70, "r": 30, "t": 110, "b": 60},
    )
    figure.update_xaxes(title_text="Month")
    figure.update_yaxes(title_text="Index")
    return figure


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a fixed-basket high-frequency leading index for Australia from local ABS household spending workbooks."
    )
    parser.add_argument(
        "--household-spending-dir",
        type=Path,
        default=Path("abs_workbooks/monthly_household_spending"),
        help="Directory containing the ABS monthly household spending XLSX files.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("high_frequency_leading_index.html"),
        help="Output HTML path for the Plotly dashboard.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    spending_dir = args.household_spending_dir.resolve()
    if not spending_dir.exists():
        raise FileNotFoundError(f"Household spending directory not found: {spending_dir}")

    workbook_path = spending_dir / "5682002.xlsx"
    series_by_key = load_series(workbook_path)

    transformed = {key: compute_standardized_growth_gap(series) for key, series in series_by_key.items()}
    component_panel = build_balanced_panel(
        {series_by_key[key].name: transformed[key] for key in SERIES_SPECS_BY_KEY}
    )
    composite = compute_diffusion_depth_index(component_panel)
    composite.name = "Composite"

    composite_momentum = composite.diff(3).dropna()
    if composite_momentum.empty:
        raise ValueError("Composite momentum is empty after calculating the 3-month change.")

    figure = create_figure(composite)

    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.write_html(output_path, include_plotlyjs=True, full_html=True)

    print(f"Wrote {output_path}")
    print(
        f"Sample starts {component_panel.index.min().date().isoformat()} and ends {component_panel.index.max().date().isoformat()}"
    )
    print(
        f"Latest composite = {composite.iloc[-1]:.2f}; latest 3-month composite change = {composite_momentum.iloc[-1]:.2f}"
    )


SERIES_SPECS_BY_KEY = {spec["key"]: spec for spec in SERIES_SPECS}


if __name__ == "__main__":
    main()
